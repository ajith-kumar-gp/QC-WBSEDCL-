"""
Meter OCR Annotation Server  —  Production-ready version
=========================================================
Run:          python server.py
Annotator:    http://localhost:5000
Dashboard:    http://localhost:5000/dashboard
Exports:      ./exports/

WHAT SURVIVES A FULL SHUTDOWN:
  ✅ All annotations      → annotations.db  (SQLite on disk)
  ✅ Completed Excels     → exports/*.xlsx  (files on disk)
  ✅ Same image sample    → RANDOM_SEED=42  (deterministic every restart)
  ✅ User selection       → browser localStorage (per browser)
  ✅ Agency progress      → rebuilt from DB on first request

PERFORMANCE DESIGN:
  • UI boots with only agency metadata (~2 KB) — NOT 64k rows
  • Rows are fetched per-agency on demand and cached in the browser
  • Server caches sampled rows in memory after first load
  • SQLite runs in WAL mode (concurrent reads + 1 write, no blocking)
  • DB indexes on agency + row_id for fast lookups
  • Gzip compression on all API responses
  • Thread-safe DB connections (one per request)
"""

import os, hashlib, threading, time, gzip
from functools import wraps
import datetime

from flask import Flask, jsonify, request, send_from_directory, send_file, Response
import pandas as pd
import psycopg2
from psycopg2.extras import RealDictCursor
import os
from dotenv import load_dotenv
load_dotenv()

DB_CONFIG = {
    "host": os.getenv("DB_HOST"),
    "port": os.getenv("DB_PORT"),
    "database": os.getenv("DB_NAME"),
    "user": os.getenv("DB_USER"),
    "password": os.getenv("DB_PASSWORD"),
}

# ─────────────────────────────────────────────────────────
#  CONFIG  —  edit these
# ─────────────────────────────────────────────────────────
DATA_FILE         = "data.xlsx"      # or .xlsx
AGENCY_COLUMN     = "Agency"        # column name for agency/division
SAMPLE_PER_AGENCY = 150             # 150 or 200 images per agency
RANDOM_SEED       = 42              # never change this — keeps sample identical across restarts
DB_FILE           = "annotations.db"
EXPORTS_DIR       = "exports"
PORT              = 5000
HOST              = "0.0.0.0"

USERS = [
    {"id": f"u{i}", "name": f"User {i}", "color": "#5b8dee"}
    for i in range(1, 8)
]
# ─────────────────────────────────────────────────────────

app = Flask(__name__, static_folder=".", static_url_path="")
os.makedirs(EXPORTS_DIR, exist_ok=True)

# ─── GZIP MIDDLEWARE ──────────────────────────────────────
def gzip_response(f):
    """Decorator: gzip any JSON response if client accepts it."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        rv = f(*args, **kwargs)
        # only compress Response objects with JSON content
        if not isinstance(rv, Response):
            rv = app.make_response(rv)
        if (rv.content_type and 'json' in rv.content_type and
                'gzip' in request.headers.get('Accept-Encoding', '')):
            compressed = gzip.compress(rv.get_data())
            rv.set_data(compressed)
            rv.headers['Content-Encoding'] = 'gzip'
            rv.headers['Content-Length']   = len(compressed)
        return rv
    return wrapper

# ─── DATABASE ─────────────────────────────────────────────
# SQLite in WAL mode: multiple readers + 1 writer, no blocking.
# Each request gets its own connection (thread-safe).

def get_db():
    conn = psycopg2.connect(**DB_CONFIG)
    return conn

def init_db():

    conn = get_db()
    cur = conn.cursor()

    # annotations table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS annotations (
            id SERIAL PRIMARY KEY,
            row_id TEXT UNIQUE,
            agency TEXT,
            consumer_id TEXT,
            img_url TEXT,
            label TEXT,
            notes TEXT,
            skipped BOOLEAN DEFAULT FALSE,
            annotated_by TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)

    # completed agencies table
    cur.execute("""
        CREATE TABLE IF NOT EXISTS completed_agencies (
            agency TEXT PRIMARY KEY,
            completed_at TIMESTAMP,
            excel_path TEXT,
            total_rows INTEGER
        )
    """)

    conn.commit()

    cur.close()
    conn.close()

    print("[db] PostgreSQL tables ready")
# ─── DATA LOADING ─────────────────────────────────────────
# Loaded once at startup, stays in memory for the server's lifetime.
# On restart the same rows are produced (RANDOM_SEED is fixed).

_lock         = threading.Lock()
_rows_cache   = None   # flat list of all sampled rows
_agency_index = None   # dict: agency_name -> [row, ...]
_agency_names = None   # sorted list of agency names
_total_rows   = 0

def load_and_sample():
    global _rows_cache, _agency_index, _agency_names, _total_rows
    with _lock:
        if _rows_cache is not None:
            return _rows_cache

        t0 = time.time()
        print(f"[data] Loading {DATA_FILE} ...")
        ext = os.path.splitext(DATA_FILE)[1].lower()
        df  = (pd.read_excel(DATA_FILE, engine="openpyxl")
               if ext in (".xlsx", ".xls") else pd.read_csv(DATA_FILE))
        df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]

        rename_map = {
            "actual_reading": "actual_reading",
            "actualreading": "actual_reading",

            "ai_based_meter_reading": "ai_meter_reading",
            "ai_meter_reading": "ai_meter_reading",

            "ai_based_confidence_level": "confidence_score",
            "confidence_score": "confidence_score",

            "consumerid": "consumer_id",
        }

        df.rename(columns=rename_map, inplace=True)

        for old, new in rename_map.items():
            if old in df.columns:
                df.rename(columns={old: new}, inplace=True)

        agency_col = next((c for c in df.columns if c.lower() == AGENCY_COLUMN.lower()), None)
        if not agency_col:
            raise ValueError(
                f"Agency column '{AGENCY_COLUMN}' not found.\n"
                f"Available columns: {list(df.columns)}"
            )

        parts = []
        for agency, group in df.groupby(agency_col, sort=True):
            if len(group) <= SAMPLE_PER_AGENCY:
                parts.append(group)
            else:
                parts.append(group.sample(n=SAMPLE_PER_AGENCY, random_state=RANDOM_SEED))

        sampled = pd.concat(parts, ignore_index=True)
        n_agencies = sampled[agency_col].nunique()
        print(f"[data] {len(sampled):,} rows across {n_agencies} agencies "
              f"({time.time()-t0:.1f}s)")

        rows, idx = [], {}
        for _, row in sampled.iterrows():
            d = {
                k: (
                    None if pd.isna(v)
                    else str(v) if isinstance(v, (datetime.datetime, datetime.date, datetime.time, pd.Timestamp))
                    else v
                )
                for k, v in row.items()
            }
            # Stable row ID — same every restart for the same data
            raw = (f"{d.get('consumer_id','')}"
                   f"|{d.get('img_url', d.get('Img_url', d.get('IMG_URL', d.get('image_url',''))))}")
            d["row_id"]  = hashlib.md5(raw.encode()).hexdigest()
            d["_agency"] = str(d.get(agency_col) or "Unknown")
            rows.append(d)
            idx.setdefault(d["_agency"], []).append(d)

        _rows_cache   = rows
        _agency_index = idx
        _agency_names = sorted(idx.keys())
        _total_rows   = len(rows)

        assign_agencies_to_users()   # <-- ADD THIS LINE

        print(f"[data] Ready — {_total_rows:,} rows cached in memory")
        return rows
    
USER_ASSIGNMENTS = {}

def assign_agencies_to_users():

    global USER_ASSIGNMENTS

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT DISTINCT agency
        FROM meter_images
        ORDER BY agency
    """)

    agencies = [r[0] for r in cur.fetchall()]

    cur.close()
    conn.close()

    users = [u["id"] for u in USERS]

    USER_ASSIGNMENTS = {u: [] for u in users}

    for i, agency in enumerate(agencies):
        user = users[i % len(users)]
        USER_ASSIGNMENTS[user].append(agency)

    print("[server] Agency assignments created") 
def get_agency_rows(agency: str):
    load_and_sample()
    return _agency_index.get(agency, [])

def get_agency_names():
    load_and_sample()
    return _agency_names

# ─── AGENCY STATS HELPER ──────────────────────────────────
# Called often — uses indexed DB queries, never touches the full row cache.

def get_agency_stats(db):
    """Returns dict: agency -> {annotated, by_user, labels}"""
    cur = db.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
    SELECT agency, label, skipped, annotated_by, COUNT(*) as cnt
    FROM annotations
    GROUP BY agency, label, skipped, annotated_by
    """)

    rows = cur.fetchall()

    result = {}
    for r in rows:
        ag = r["agency"]
        if ag not in result:
            result[ag] = {"annotated": 0, "by_user": {}, "labels": {}}
        result[ag]["annotated"] += r["cnt"]
        uid = r["annotated_by"] or "unknown"
        result[ag]["by_user"][uid] = result[ag]["by_user"].get(uid, 0) + r["cnt"]
        lbl = "skipped" if r["skipped"] else (r["label"] or "unlabeled")
        result[ag]["labels"][lbl] = result[ag]["labels"].get(lbl, 0) + r["cnt"]
    return result

# ─── EXCEL BUILDER ────────────────────────────────────────

def build_agency_excel(agency, agency_rows, ann_map, user_map):
    safe      = "".join(c if c.isalnum() or c in " ._-" else "_" for c in agency)
    ts        = datetime.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename  = f"{safe}_{ts}.xlsx"
    filepath  = os.path.join(EXPORTS_DIR, filename)

    src_keys  = [k for k in (agency_rows[0].keys() if agency_rows else [])
                 if not k.startswith("_") and k != "row_id"]
    ann_cols  = ["annotation_label","annotation_notes","skipped",
                 "annotated_by_id","annotated_by_name",
                 "annotated_at","created_at","row_id"]

    records = []
    for row in agency_rows:
        ann   = ann_map.get(row["row_id"], {})
        uid   = ann.get("annotated_by","")
        rec   = {k: ("" if row.get(k) is None else row.get(k)) for k in src_keys}
        rec.update({
            "annotation_label":  ann.get("label",""),
            "annotation_notes":  ann.get("notes",""),
            "skipped":           "Yes" if ann.get("skipped") else "No",
            "annotated_by_id":   uid,
            "annotated_by_name": user_map.get(uid, uid),
            "annotated_at":      ann.get("updated_at",""),
            "created_at":        ann.get("created_at",""),
            "row_id":            row["row_id"],
        })
        records.append(rec)

    df_all = pd.DataFrame(records, columns=src_keys + ann_cols)

    lc = df_all["annotation_label"].value_counts().reset_index()
    lc.columns = ["Label","Count"]
    lc["Pct"] = (lc["Count"] / max(len(df_all),1) * 100).round(1).astype(str) + "%"

    by_ann = (df_all.groupby("annotated_by_name")
                    .agg(Count=("row_id","count"))
                    .reset_index()
                    .rename(columns={"annotated_by_name":"Annotator"}))

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="All Rows", index=False)

        meta = pd.DataFrame([
            ["Agency",       agency],
            ["Generated",    datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")],
            ["Total Rows",   len(df_all)],
            ["Correct",      int((df_all["annotation_label"]=="actual_correct").sum())],
            ["Wrong",        int((df_all["annotation_label"]=="actual_wrong").sum())],
            ["Skipped",      int((df_all["skipped"]=="Yes").sum())],
            ["Other",        int(~df_all["annotation_label"].isin(
                                  ["actual_correct","actual_wrong",""]).sum())],
        ], columns=["Field","Value"])
        meta.to_excel(writer, sheet_name="Summary", index=False, startrow=0)
        lc.to_excel(writer, sheet_name="Summary", index=False, startrow=len(meta)+2)
        by_ann.to_excel(writer, sheet_name="Summary", index=False,
                        startrow=len(meta)+len(lc)+5)

        df_all[df_all["annotation_label"]=="actual_correct"].to_excel(
            writer, sheet_name="Correct", index=False)
        df_all[df_all["annotation_label"]=="actual_wrong"].to_excel(
            writer, sheet_name="Wrong", index=False)
        df_all[~df_all["annotation_label"].isin(["actual_correct","actual_wrong"]) |
               (df_all["skipped"]=="Yes")].to_excel(
            writer, sheet_name="Other", index=False)

        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            def style_ws(ws, hdr_hex, font_hex="F5A623"):
                fill  = PatternFill("solid", fgColor=hdr_hex)
                font  = Font(bold=True, color=font_hex)
                align = Alignment(horizontal="center", vertical="center")
                for cell in ws[1]:
                    cell.fill = fill; cell.font = font; cell.alignment = align
                ws.freeze_panes = "A2"
                for i, col in enumerate(ws.columns, 1):
                    w = max((len(str(c.value or "")) for c in col), default=8)
                    ws.column_dimensions[get_column_letter(i)].width = min(w + 4, 44)

            style_ws(writer.sheets["All Rows"], "161719")
            style_ws(writer.sheets["Summary"],  "161719")
            style_ws(writer.sheets["Correct"],  "1B4332", "FFFFFF")
            style_ws(writer.sheets["Wrong"],    "4B1C1C", "FFFFFF")
            style_ws(writer.sheets["Other"],    "1A1A2E")
        except Exception as e:
            print(f"[excel] Styling skipped: {e}")

    print(f"[excel] Saved → {filepath}")
    return filepath, filename

# ─── COMPLETION CHECK ─────────────────────────────────────

def check_and_complete_agency(agency, db):
    """Returns (just_completed, filename_or_None)"""
    if db.execute("SELECT 1 FROM completed_agencies WHERE agency=%s", (agency,)).fetchone():
        return False, None   # already done

    total_needed = len(get_agency_rows(agency))
    if total_needed == 0:
        return False, None

    ann_count = db.execute(
        "SELECT COUNT(*) FROM annotations WHERE agency=%s", (agency,)
    ).fetchone()[0]

    if ann_count < total_needed:
        return False, None

    # 🎉 Just finished this agency
    ann_map  = {r["row_id"]: dict(r) for r in
                db.execute("SELECT * FROM annotations WHERE agency=%s", (agency,)).fetchall()}
    user_map = {u["id"]: u["name"] for u in USERS}
    try:
        filepath, filename = build_agency_excel(
            agency, get_agency_rows(agency), ann_map, user_map)
        db.execute(
            "INSERT INTO completed_agencies (agency,completed_at,excel_path,total_rows) "
            "VALUES (%s,%s,%s,%s)",
            (agency, datetime.datetime.utcnow().isoformat(), filepath, total_needed))
        db.commit()
        return True, filename
    except Exception as e:
        print(f"[server] Excel build failed for '{agency}': {e}")
        return False, None

# ─── ROUTES ───────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(".", "ui.html")

@app.route("/dashboard")
def dashboard():
    return send_from_directory(".", "dashboard.html")

# ── Static data (users never change at runtime)
@app.route("/api/users")
def api_users():
    resp = jsonify(USERS)
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp

# ── Agency list  (lightweight — no row data, just names + counts)
# This is what the UI fetches on boot. Tiny payload.
@app.route("/api/agencies")
def api_agencies():
    print("[API] /api/agencies called")

    user = request.args.get("user")

    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT agency, COUNT(*) as total
        FROM meter_images
        GROUP BY agency
        ORDER BY agency
    """)

    agencies = cur.fetchall()

    cur.execute("""
        SELECT agency, COUNT(*) as annotated
        FROM annotations
        GROUP BY agency
    """)

    ann_counts = {r["agency"]: r["annotated"] for r in cur.fetchall()}

    allowed_agencies = USER_ASSIGNMENTS.get(user, []) if user else []

    result = []

    for row in agencies:

        if user and row["agency"] not in allowed_agencies:
            continue

        annotated = ann_counts.get(row["agency"], 0)

        result.append({
            "agency": row["agency"],
            "total": row["total"],
            "annotated": annotated,
            "pending": row["total"] - annotated,
            "completed": annotated >= row["total"]
        })

    cur.close()
    conn.close()

    return jsonify(result)

@app.route("/api/images/<path:agency>")
def api_images_for_agency(agency):

    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT
            row_id,
            consumer_id,
            actual_reading,
            ai_meter_reading,
            confidence_score,
            img_url,
            agency
        FROM meter_images
        WHERE agency=%s
        LIMIT 150
    """, (agency,))

    rows = cur.fetchall()

    cur.execute("""
        SELECT row_id,label,notes,skipped
        FROM annotations
        WHERE agency=%s
    """, (agency,))

    annotations = {r["row_id"]: r for r in cur.fetchall()}

    result = []

    for r in rows:

        ann = annotations.get(r["row_id"], {})

        r["_saved_label"] = ann.get("label", "")
        r["_saved_notes"] = ann.get("notes", "")
        r["_saved_skipped"] = ann.get("skipped", False)

        result.append(r)

    cur.close()
    conn.close()

    return jsonify(result)
# ── Kept for backward-compat / bulk export  (not used by UI boot)
@app.route("/api/images")
@gzip_response
def api_images_all():
    agency = request.args.get("agency")
    if agency:
        return api_images_for_agency(agency)
    rows = load_and_sample()
    return jsonify([{k:v for k,v in r.items() if not k.startswith("_")} for r in rows])

# ── Save annotation
@app.route("/api/annotate", methods=["POST"])
def api_annotate():
    print("[server] Annotation received")
    data = request.get_json()

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
    INSERT INTO annotations (
        row_id,
        consumer_id,
        agency,
        img_url,
        label,
        notes,
        skipped,
        annotated_by
    )
    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
    ON CONFLICT (row_id)
    DO UPDATE SET
        label = EXCLUDED.label,
        notes = EXCLUDED.notes,
        skipped = EXCLUDED.skipped,
        annotated_by = EXCLUDED.annotated_by,
        created_at = NOW()
    """, (
        data.get("row_id"),
        data.get("consumer_id"),
        data.get("agency"),
        data.get("img_url"),
        data.get("label"),
        data.get("notes"),
        data.get("skipped"),
        data.get("annotated_by")
    ))

    conn.commit()

    cur.close()
    conn.close()

    return jsonify({"ok": True ,"agency_completed": False})



# ── Download completed Excel



@app.route("/api/exports/<filename>")
def api_download_export(filename):
    safe = os.path.basename(filename)
    path = os.path.join(EXPORTS_DIR, safe)
    if not os.path.exists(path):
        return jsonify({"error": "File not found"}), 404
    return send_file(path, as_attachment=True, download_name=safe,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── List completed agencies
@app.route("/api/completed_agencies")
@gzip_response
def api_completed_agencies():
    db   = get_db()
    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
    SELECT agency, completed_at, excel_path, total_rows
    FROM completed_agencies
    ORDER BY completed_at DESC
    """)

    rows = cur.fetchall()
    result = []
    for r in rows:
        fname = os.path.basename(r["excel_path"]) if r["excel_path"] else None
        result.append({
            "agency":       r["agency"],
            "completed_at": r["completed_at"],
            "total_rows":   r["total_rows"],
            "excel_url":    f"/api/exports/{fname}" if fname else None,
            "excel_name":   fname,
        })
    return jsonify(result)

# ── Force-rebuild Excel for an agency
@app.route("/api/regenerate_excel/<path:agency>", methods=["POST"])
def api_regenerate_excel(agency):
    agency_rows = get_agency_rows(agency)
    if not agency_rows:
        return jsonify({"error":"Agency not found"}), 404
    db      = get_db()
    ann_map = {r["row_id"]: dict(r) for r in
               db.execute("SELECT * FROM annotations WHERE agency=%s", (agency,)).fetchall()}
    try:
        filepath, filename = build_agency_excel(
            agency, agency_rows, ann_map, {u["id"]:u["name"] for u in USERS})
        db.execute(
            "INSERT OR REPLACE INTO completed_agencies (agency,completed_at,excel_path,total_rows)"
            "VALUES (%s,%s,%s,%s)",
            (agency, datetime.datetime.utcnow().isoformat(), filepath, len(agency_rows)))
        db.commit()
        return jsonify({"ok":True, "excel_url":f"/api/exports/{filename}", "excel_name":filename})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

# ── Annotations list (dashboard / export)
@app.route("/api/annotations")
@gzip_response
def api_annotations():

    agency = request.args.get("agency")
    user = request.args.get("user")
    limit = int(request.args.get("limit", 500))

    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    q = "SELECT * FROM annotations WHERE 1=1"
    params = []

    if agency:
        q += " AND agency=%s"
        params.append(agency)

    if user:
        q += " AND annotated_by=%s"
        params.append(user)

    q += " ORDER BY created_at DESC LIMIT %s"
    params.append(limit)

    cur.execute(q, params)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return jsonify(rows)
# ── Dashboard summary (per-user stats)
@app.route("/api/dashboard")
def api_dashboard():

    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    # total images
    cur.execute("SELECT COUNT(*) FROM meter_images")
    total_images = cur.fetchone()["count"]

    # total agencies
    cur.execute("SELECT COUNT(DISTINCT agency) FROM meter_images")
    total_agencies = cur.fetchone()["count"]

    # total annotated
    cur.execute("SELECT COUNT(*) FROM annotations")
    total_annotated = cur.fetchone()["count"]

    # completed agencies
    cur.execute("SELECT COUNT(*) FROM completed_agencies")
    completed_agencies = cur.fetchone()["count"]

    # user stats
    cur.execute("""
        SELECT
            annotated_by,
            COUNT(*) as total_annotations
        FROM annotations
        GROUP BY annotated_by
    """)

    rows = cur.fetchall()

    users = []

    for r in rows:

        uid = r["annotated_by"]

        user = next((u for u in USERS if u["id"] == uid), None)
        if not user:
            continue

        users.append({
            "id": uid,
            "name": user["name"],
            "color": user["color"],
            "total_annotations": r["total_annotations"],
            "agencies_touched": 0,
            "agencies_list": [],
            "labels": {}
        })

    cur.close()
    conn.close()

    return jsonify({
        "global": {
            "total_images": total_images,
            "total_annotated": total_annotated,
            "total_agencies": total_agencies,
            "completed_agencies": completed_agencies
        },
        "users": users
    })

@app.route("/api/export/csv")
def api_export_csv():

    import io, csv

    conn = get_db()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT 
            m.agency,
            m.consumer_id,
            m.actual_reading,
            m.ai_meter_reading,
            m.confidence_score,
            m.img_url,
            a.label,
            a.notes,
            a.skipped,
            a.annotated_by,
            a.created_at
        FROM meter_images m
        LEFT JOIN annotations a
        ON m.row_id = a.row_id
        ORDER BY m.agency
    """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    out = io.StringIO()

    if not rows:
        return Response("", mimetype="text/csv")

    writer = csv.DictWriter(out, fieldnames=rows[0].keys())
    writer.writeheader()

    for r in rows:
        writer.writerow(r)

    fname = f"annotations_{datetime.datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"

    return Response(
        out.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    ) 
# ─── STARTUP ──────────────────────────────────────────────
if __name__ == "__main__":
    init_db()
    assign_agencies_to_users()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
